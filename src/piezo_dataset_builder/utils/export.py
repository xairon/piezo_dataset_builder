"""
Utilitaires pour l'export de données.
"""

import pandas as pd
from io import BytesIO, StringIO
import logging
import zipfile
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def to_csv(df: pd.DataFrame) -> bytes:
    """
    Exporte DataFrame en CSV (bytes).

    Args:
        df: DataFrame à exporter

    Returns:
        Bytes du CSV encodé en UTF-8
    """
    try:
        csv_data = df.to_csv(index=False).encode('utf-8')
        logger.info(f"Exported CSV: {len(df)} rows, {len(df.columns)} columns")
        return csv_data
    except Exception as e:
        logger.error(f"Error exporting to CSV: {e}")
        raise


def to_excel(df: pd.DataFrame, sheet_name: str = 'Dataset') -> bytes:
    """
    Exporte DataFrame en Excel (bytes) avec auto-ajustement des colonnes.

    Args:
        df: DataFrame à exporter
        sheet_name: Nom de la feuille Excel

    Returns:
        Bytes du fichier Excel
    """
    buffer = BytesIO()

    try:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Auto-ajuster largeur des colonnes
            worksheet = writer.sheets[sheet_name]

            for col_idx, column in enumerate(df.columns):
                # Calculate column width
                column_length = max(
                    df[column].astype(str).map(len).max(),
                    len(str(column))
                )
                # Limite max pour éviter des colonnes trop larges
                column_length = min(column_length, 50)

                # Use openpyxl.utils.get_column_letter for correct Excel column naming
                # Handles columns beyond Z (AA, AB, etc.)
                col_letter = get_column_letter(col_idx + 1)
                worksheet.column_dimensions[col_letter].width = column_length + 2

        logger.info(
            f"Exported Excel: {len(df)} rows, {len(df.columns)} columns, "
            f"sheet='{sheet_name}'"
        )
        return buffer.getvalue()

    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise


def to_json(df: pd.DataFrame, orient: str = 'records') -> str:
    """
    Exporte DataFrame en JSON.

    Args:
        df: DataFrame à exporter
        orient: Format JSON ('records', 'split', 'table', etc.)

    Returns:
        String JSON
    """
    try:
        json_data = df.to_json(
            orient=orient,
            date_format='iso',
            indent=2,
            force_ascii=False
        )
        logger.info(
            f"Exported JSON: {len(df)} rows, {len(df.columns)} columns, "
            f"orient='{orient}'"
        )
        return json_data

    except Exception as e:
        logger.error(f"Error exporting to JSON: {e}")
        raise


def get_export_stats(df: pd.DataFrame) -> dict:
    """
    Calcule les statistiques du dataset pour l'export.

    Args:
        df: DataFrame

    Returns:
        Dict avec statistiques
    """
    if df.empty:
        logger.warning("get_export_stats called on empty DataFrame")
        return {
            'nb_lignes': 0,
            'nb_colonnes': 0,
            'taille_mo': 0.0,
        }

    stats = {
        'nb_lignes': len(df),
        'nb_colonnes': len(df.columns),
        'taille_mo': df.memory_usage(deep=True).sum() / 1024 / 1024,  # En Mo
    }

    # Stats par type de colonne
    # Supporte code_bss (standard) ou code_station (legacy/meteo)
    station_col = 'code_bss' if 'code_bss' in df.columns else 'code_station'
    if station_col in df.columns:
        stats['nb_stations'] = df[station_col].nunique()

    if 'date' in df.columns:
        try:
            dates = pd.to_datetime(df['date'], errors='coerce')
            # Remove NaT values
            dates = dates.dropna()
            if not dates.empty:
                stats['date_min'] = dates.min()
                stats['date_max'] = dates.max()
                stats['nb_jours'] = (stats['date_max'] - stats['date_min']).days + 1
        except Exception as e:
            logger.debug(f"Could not compute date stats: {e}")

    # Taux de valeurs manquantes
    total_cells = len(df) * len(df.columns)
    if total_cells > 0:
        stats['taux_na'] = (df.isna().sum().sum() / total_cells) * 100
    else:
        stats['taux_na'] = 0.0

    logger.debug(f"Export stats: {stats}")
    return stats


def to_zip_by_station(df: pd.DataFrame, file_format: str = 'csv') -> bytes:
    """
    Exporte DataFrame en archive ZIP avec un fichier par station.

    Args:
        df: DataFrame à exporter
        file_format: Format des fichiers ('csv' ou 'excel')

    Returns:
        Bytes du fichier ZIP
    """
    # Déterminer la colonne station
    station_col = 'code_bss' if 'code_bss' in df.columns else 'code_station'

    if station_col not in df.columns:
        raise ValueError(f"No station column found (code_bss or code_station)")

    buffer = BytesIO()

    try:
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            stations = df[station_col].unique()

            for station in stations:
                df_station = df[df[station_col] == station]

                # Nettoyer le nom de fichier (remplacer caractères invalides)
                safe_name = str(station).replace('/', '_').replace('\\', '_')

                if file_format == 'csv':
                    content = df_station.to_csv(index=False).encode('utf-8')
                    filename = f"{safe_name}.csv"
                elif file_format == 'excel':
                    excel_buffer = BytesIO()
                    df_station.to_excel(excel_buffer, index=False, engine='openpyxl')
                    content = excel_buffer.getvalue()
                    filename = f"{safe_name}.xlsx"
                else:
                    raise ValueError(f"Unknown format: {file_format}")

                zf.writestr(filename, content)

            logger.info(f"Exported ZIP archive: {len(stations)} station files ({file_format})")

        return buffer.getvalue()

    except Exception as e:
        logger.error(f"Error exporting to ZIP: {e}")
        raise
